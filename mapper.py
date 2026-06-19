#!/usr/bin/env python3
"""NIS2 → ISO 27001:2022 Gap-Mapper CLI"""

import argparse
import sys
from datetime import date
from pathlib import Path

import openpyxl
from jinja2 import Environment, FileSystemLoader
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from mapping_data import CONTROL_NAMES, NIS2_MAPPING

STATUS_UMGESETZT = "Umgesetzt"
STATUS_TEILWEISE = "Teilweise"
STATUS_NICHT = "Nicht umgesetzt"
STATUS_UNBEKANNT = "Unbekannt"

COLOR_GREEN = "FF92D050"
COLOR_YELLOW = "FFFFC000"
COLOR_RED = "FFFF0000"
COLOR_GREY = "FFD9D9D9"


def read_controls(path: Path) -> dict[str, dict]:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    col = {name: idx for idx, name in enumerate(headers)}
    required = {"control_id", "status"}
    missing = required - col.keys()
    if missing:
        sys.exit(f"Fehler: Spalten fehlen in der Eingabedatei: {missing}")

    controls = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[col["control_id"]]:
            continue
        cid = str(row[col["control_id"]]).strip()
        status = str(row[col["status"]]).strip() if row[col["status"]] else STATUS_UNBEKANNT
        name = str(row[col.get("control_name", -1)]).strip() if col.get("control_name") is not None and row[col.get("control_name")] else CONTROL_NAMES.get(cid, "")
        notes = str(row[col.get("notes", -1)]).strip() if col.get("notes") is not None and row[col.get("notes")] else ""
        responsible = str(row[col.get("responsible", -1)]).strip() if col.get("responsible") is not None and row[col.get("responsible")] else ""
        controls[cid] = {
            "control_id": cid,
            "control_name": name,
            "status": status,
            "notes": notes,
            "responsible": responsible,
        }
    return controls


def compute_nis2_status(nis2_req: dict, controls: dict[str, dict]) -> dict:
    mapped = []
    for cid in nis2_req["controls"]:
        ctrl = controls.get(cid, {
            "control_id": cid,
            "control_name": CONTROL_NAMES.get(cid, ""),
            "status": STATUS_UNBEKANNT,
            "notes": "",
            "responsible": "",
        })
        mapped.append(ctrl)

    statuses = [c["status"] for c in mapped]
    umgesetzt = sum(1 for s in statuses if s == STATUS_UMGESETZT)
    teilweise = sum(1 for s in statuses if s == STATUS_TEILWEISE)

    if umgesetzt == len(mapped):
        rag = "Erfüllt"
    elif umgesetzt > 0 or teilweise > 0:
        rag = "Teilweise"
    else:
        rag = "Lücke"

    return {
        "controls": mapped,
        "rag": rag,
        "umgesetzt": umgesetzt,
        "teilweise": teilweise,
        "total": len(mapped),
    }


def build_report_data(controls: dict[str, dict]) -> dict:
    results = {}
    for req_name, req_data in NIS2_MAPPING.items():
        results[req_name] = {
            "beschreibung": req_data["beschreibung"],
            **compute_nis2_status(req_data, controls),
        }

    total = len(results)
    erfuellt = sum(1 for r in results.values() if r["rag"] == "Erfüllt")
    teilweise_count = sum(1 for r in results.values() if r["rag"] == "Teilweise")
    luecken = sum(1 for r in results.values() if r["rag"] == "Lücke")
    quote = round(erfuellt / total * 100) if total else 0

    return {
        "results": results,
        "gesamt_quote": quote,
        "gesamt_erfuellt": erfuellt,
        "gesamt_teilweise": teilweise_count,
        "gesamt_luecken": luecken,
        "gesamt_total": total,
        "datum": date.today().strftime("%d.%m.%Y"),
    }


def write_html(data: dict, output_dir: Path, template_dir: Path) -> Path:
    env = Environment(loader=FileSystemLoader(str(template_dir)), autoescape=True)
    tmpl = env.get_template("report.html.j2")
    html = tmpl.render(**data)
    out = output_dir / "gap_report.html"
    out.write_text(html, encoding="utf-8")
    return out


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _header_font() -> Font:
    return Font(bold=True, color="FFFFFFFF")


def _header_fill() -> PatternFill:
    return PatternFill(start_color="FF2E4057", end_color="FF2E4057", fill_type="solid")


def _rag_color(rag: str) -> str:
    return {
        "Erfüllt": COLOR_GREEN,
        "Teilweise": COLOR_YELLOW,
        "Lücke": COLOR_RED,
    }.get(rag, COLOR_GREY)


def _status_color(status: str) -> str:
    return {
        STATUS_UMGESETZT: COLOR_GREEN,
        STATUS_TEILWEISE: COLOR_YELLOW,
        STATUS_NICHT: COLOR_RED,
    }.get(status, COLOR_GREY)


def write_excel(data: dict, output_dir: Path) -> Path:
    wb = openpyxl.Workbook()

    # ── Sheet 1: Übersicht ──────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "NIS2 Übersicht"

    ws_sum["A1"] = "NIS2 → ISO 27001:2022 Gap-Report"
    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum["A2"] = f"Erstellt am: {data['datum']}"
    ws_sum["A3"] = f"Gesamterfüllungsquote: {data['gesamt_quote']} %"
    ws_sum["A3"].font = Font(bold=True)

    ws_sum.append([])
    headers = ["NIS2-Anforderung", "Beschreibung", "Status", "Erfüllt", "Teilweise", "Lücken", "Gesamt"]
    ws_sum.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws_sum.cell(row=5, column=col_idx)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for req_name, result in data["results"].items():
        row = [
            req_name,
            result["beschreibung"],
            result["rag"],
            result["umgesetzt"],
            result["teilweise"],
            result["gesamt_luecken"] if "gesamt_luecken" in result else result["total"] - result["umgesetzt"] - result["teilweise"],
            result["total"],
        ]
        ws_sum.append(row)
        r = ws_sum.max_row
        ws_sum.cell(r, 3).fill = _fill(_rag_color(result["rag"]))
        ws_sum.cell(r, 3).alignment = Alignment(horizontal="center")
        ws_sum.cell(r, 1).alignment = Alignment(wrap_text=True)
        ws_sum.cell(r, 2).alignment = Alignment(wrap_text=True)

    ws_sum.column_dimensions["A"].width = 45
    ws_sum.column_dimensions["B"].width = 55
    ws_sum.column_dimensions["C"].width = 14
    for i in range(4, 8):
        ws_sum.column_dimensions[get_column_letter(i)].width = 12

    # ── Sheet 2: Detail ─────────────────────────────────────────────────────
    ws_det = wb.create_sheet("Control-Details")
    det_headers = ["NIS2-Anforderung", "Control-ID", "Control-Name", "Status", "Verantwortlich", "Notizen"]
    ws_det.append(det_headers)
    for col_idx, _ in enumerate(det_headers, start=1):
        cell = ws_det.cell(row=1, column=col_idx)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for req_name, result in data["results"].items():
        for ctrl in result["controls"]:
            row = [
                req_name,
                ctrl["control_id"],
                ctrl["control_name"],
                ctrl["status"],
                ctrl["responsible"],
                ctrl["notes"],
            ]
            ws_det.append(row)
            r = ws_det.max_row
            ws_det.cell(r, 4).fill = _fill(_status_color(ctrl["status"]))
            ws_det.cell(r, 4).alignment = Alignment(horizontal="center")
            ws_det.cell(r, 1).alignment = Alignment(wrap_text=True)
            ws_det.cell(r, 3).alignment = Alignment(wrap_text=True)
            ws_det.cell(r, 6).alignment = Alignment(wrap_text=True)

    ws_det.column_dimensions["A"].width = 45
    ws_det.column_dimensions["B"].width = 12
    ws_det.column_dimensions["C"].width = 50
    ws_det.column_dimensions["D"].width = 18
    ws_det.column_dimensions["E"].width = 20
    ws_det.column_dimensions["F"].width = 40

    out = output_dir / "gap_report.xlsx"
    wb.save(out)
    return out


def main():
    parser = argparse.ArgumentParser(
        description="NIS2 Art. 21 → ISO 27001:2022 Gap-Mapper",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Beispiel:
  python mapper.py --input sample_input/controls_sample.xlsx --output output/
        """,
    )
    parser.add_argument("--input", required=True, help="Pfad zur controls_status.xlsx")
    parser.add_argument("--output", default="output/", help="Ausgabeverzeichnis (default: output/)")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        sys.exit(f"Fehler: Eingabedatei nicht gefunden: {input_path}")

    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)

    template_dir = Path(__file__).parent / "templates"

    print(f"  Lese Controls aus: {input_path}")
    controls = read_controls(input_path)
    print(f"  {len(controls)} Controls geladen.")

    data = build_report_data(controls)

    print(f"  Gesamterfüllungsquote: {data['gesamt_quote']} %  "
          f"({data['gesamt_erfuellt']} Erfüllt / {data['gesamt_teilweise']} Teilweise / {data['gesamt_luecken']} Lücke)")

    html_path = write_html(data, output_dir, template_dir)
    print(f"  HTML-Report: {html_path}")

    xlsx_path = write_excel(data, output_dir)
    print(f"  Excel-Report: {xlsx_path}")

    print("\nFertig.")


if __name__ == "__main__":
    main()
