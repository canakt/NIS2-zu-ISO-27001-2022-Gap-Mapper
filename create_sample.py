"""Hilfsskript – erzeugt sample_input/controls_sample.xlsx"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

SAMPLE_CONTROLS = [
    ("5.1",  "Richtlinien für Informationssicherheit",                         "Umgesetzt",        "ISMS-Policy v2.3 genehmigt",       "CISO"),
    ("5.2",  "Rollen und Verantwortlichkeiten für IS",                          "Umgesetzt",        "Organigramm aktualisiert",         "CISO"),
    ("5.3",  "Aufgabentrennung",                                                "Teilweise",        "Noch nicht vollständig dokumentiert", "IT-Leitung"),
    ("5.24", "Planung und Vorbereitung des IS-Incident-Managements",            "Umgesetzt",        "IR-Plan v1.5",                     "SOC"),
    ("5.25", "Beurteilung und Entscheidung über IS-Ereignisse",                 "Teilweise",        "Klassifizierungsschema in Arbeit", "SOC"),
    ("5.26", "Reaktion auf IS-Vorfälle",                                        "Nicht umgesetzt",  "",                                 ""),
    ("5.29", "IS während Unterbrechungen",                                      "Umgesetzt",        "BCP v3.0 vorhanden",               "BCM-Team"),
    ("5.30", "IKT-Bereitschaft für Business Continuity",                        "Teilweise",        "RTO/RPO definiert, Tests ausstehend", "IT-Betrieb"),
    ("8.13", "Datensicherung",                                                  "Umgesetzt",        "Täglich, 30 Tage Retention",       "IT-Betrieb"),
    ("8.14", "Redundanz von informationsverarbeitenden Einrichtungen",           "Nicht umgesetzt",  "Kein DR-Standort vorhanden",       ""),
    ("5.19", "Informationssicherheit in Lieferantenbeziehungen",                "Teilweise",        "Muster-NDA vorhanden, kein Audit", "Einkauf"),
    ("8.24", "Einsatz von Kryptographie",                                       "Umgesetzt",        "Crypto-Policy v1.0",               "IT-Security"),
    ("6.3",  "Sensibilisierung, Ausbildung und Schulung zur IS",                "Umgesetzt",        "Jährliches Awareness-Training",    "HR"),
    ("8.25", "Sicherer Entwicklungslebenszyklus",                               "Nicht umgesetzt",  "SDL noch nicht eingeführt",        ""),
    ("8.5",  "Sichere Authentifizierung",                                       "Teilweise",        "MFA für Admin-Konten aktiv",       "IAM-Team"),
    ("5.35", "Unabhängige Überprüfung der IS",                                  "Nicht umgesetzt",  "Internes Audit geplant Q3",        ""),
    ("9.1",  "Überwachung, Messung, Analyse und Bewertung",                     "Teilweise",        "KPIs definiert, Reporting lückenhaft", "CISO"),
]

out = Path("sample_input/controls_sample.xlsx")
out.parent.mkdir(parents=True, exist_ok=True)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Controls"

headers = ["control_id", "control_name", "status", "notes", "responsible"]
ws.append(headers)
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(1, col_idx)
    cell.font = Font(bold=True, color="FFFFFFFF")
    cell.fill = PatternFill(start_color="FF2E4057", end_color="FF2E4057", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

STATUS_COLORS = {
    "Umgesetzt":       "FFE8F5E9",
    "Teilweise":       "FFFFF3E0",
    "Nicht umgesetzt": "FFFFEBEE",
}

for row_data in SAMPLE_CONTROLS:
    ws.append(list(row_data))
    r = ws.max_row
    color = STATUS_COLORS.get(row_data[2], "FFD9D9D9")
    ws.cell(r, 3).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    ws.cell(r, 3).alignment = Alignment(horizontal="center")

ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 58
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 42
ws.column_dimensions["E"].width = 18

wb.save(out)
print(f"Beispieldatei erstellt: {out}")
